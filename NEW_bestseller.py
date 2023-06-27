import requests
import openpyxl
import json
from time import sleep
res = requests.get("https://athena.eslite.com/api/v1/best_sellers/online/month?l1=3&page=1&per_page=100")
jd = json.loads(res.text)
i = 0
wb = openpyxl.Workbook()
ws = wb.active

ws["A1"] = "書籍分類"             
ws["B1"] = "ISBN"
ws["C1"] = "書名"
ws["D1"] = "作者"
ws["E1"] = "出版社"
ws["F1"] = "出版日期"
ws["G1"] = "原價"
ws["H1"] = "折數"
ws["I1"] = "打折後價格"

for books in jd["products"]:
    A = " "
    B = (jd["products"][i]["ean"])
    C = (jd["products"][i]["name"])
    D = (jd["products"][i]["author"])
    E = (jd["products"][i]["manufacturer"])
    F = (jd["products"][i]["manufacturer_date"])
    G = (jd["products"][i]["final_price"])
    H = (jd["products"][i]["discount_range"])
    I = (jd["products"][i]["retail_price"])
    print("----------第"+str(i+1)+"本-----------")

    ws.append([A,B,C,D,E,F,G,H,I])
        
    i += 1
    res = requests.get("https://athena.eslite.com/api/v1/best_sellers/online/month?l1=3&page=1&per_page=100")
    jd = json.loads(res.text)
    sleep(0.1)
    wb.save("NEW誠品30日排行榜.xlsx")
