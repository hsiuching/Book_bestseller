import requests
import openpyxl
import json
from bs4 import BeautifulSoup
from time import sleep

res = requests.get("https://athena.eslite.com/api/v1/best_sellers/online/month?l1=3&page=1&per_page=100")
jd = json.loads(res.text)
i = 0
wb = openpyxl.Workbook()
ws = wb.active  #open sheet

ws["A1"] = "書籍分類"             #欄位填入資料
ws["B1"] = "書名"
ws["C1"] = "作者"
ws["D1"] = "出版社"
ws["E1"] = "出版日期"
ws["F1"] = "原價"
ws["G1"] = "折數"
ws["H1"] = "打折後價格"

for books in jd["products"]:
    A = " "
    B = (jd["products"][i]["name"])
    C = (jd["products"][i]["author"])
    D = (jd["products"][i]["manufacturer"])
    E = (jd["products"][i]["manufacturer_date"])
    F = (jd["products"][i]["final_price"])
    G = (jd["products"][i]["discount_range"])
    H = (jd["products"][i]["retail_price"])
    print("----------第"+str(i+1)+"本-----------")

    ws.append([A,B,C,D,E,F,G,H])
        
    i += 1
    res = requests.get("https://athena.eslite.com/api/v1/best_sellers/online/month?l1=3&page=1&per_page=100")
    jd = json.loads(res.text)
    sleep(0.1)
    wb.save("誠品30日排行榜.xlsx")


